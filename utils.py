from openpyxl import Workbook
import datetime

def make_assemble_list_pdf():
    pass


def make_assemble_list_xls(market, shop, assemble_list):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Дата"
    ws["B1"] = datetime.datetime.now().strftime("%Y-%m-%d")
    ws["A2"] = market
    ws["B2"] = shop
    for num, row in enumerate(assemble_list):
        ws.append([str(num+1)]+row)
    wb.save(filename := f"{datetime.datetime.now().strftime('%Y.%m.%d')} {market} {shop} Assemble list.xlsx")
    return filename

